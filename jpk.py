import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Color, Fill, Font, Alignment
from tkinter import Tk
from tkinter.filedialog import askopenfilename

from datetime import date
import os

#przygotowanie pliku excel:
raport = Workbook() 
arkusz1 = raport.active
arkusz1.title = 'sprzedaż'
arkusz2 = raport.create_sheet('zakup')

# formatowanie wybranych wierszy
ft1 = Font(name='Arial',size=8, italic=True)
ft2 = Font(name='Arial',size=9, bold=True)
wt = Alignment(wrap_text=True, horizontal='center', vertical='center')
for col in range (1,27):
    arkusz1.cell(column=col, row=17).font = ft1
    arkusz1.cell(column=col, row=18).font = ft1
    arkusz1.cell(column=col, row=19).font = ft2
    arkusz1.cell(column=col, row=17).alignment = wt
    arkusz1.cell(column=col, row=18).alignment = wt
for col in range (1,11):
    arkusz2.cell(column=col, row=17).font = ft1
    arkusz2.cell(column=col, row=18).font = ft2
    arkusz2.cell(column=col, row=17).alignment = wt

Tk().withdraw()
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
#print(filename)

tree = ET.parse(filename)
root = tree.getroot()
print(root.tag)
#print('pozycja_JPK ',root.tag.find('JPK'))
liczba = root.tag.find("JPK")
początek_taga = root.tag[:liczba]
#print('początek: ', początek)

''' NAGŁÓWEK '''

lista_naglowek = ['KodFormularza: ','WariantFormularza: ','CelZlozenia: ','DataWytworzeniaJPK: ','DataOd: ','DataDo: ','NazwaSystemu: ']
for child in root:
    if child.tag == początek_taga + 'Naglowek':
        count_rows = 6
        for wiersz in child:
            #print(wiersz.tag)
            if count_rows == 9: # te 3 zmienne potrzebne do nazwy pliku .xlsx
                data = wiersz.text[:10]
                h = wiersz.text[11:13]
                m = wiersz.text[14:16]
                s = wiersz.text[17:19]
                wersja_jpk = data + "_" + h + m + s
            if count_rows == 10:
                od = wiersz.text
            if count_rows == 11:
                do = wiersz.text
            if count_rows < 13: # wstawia informacje do kategorii nagłówka z listy lista_naglowek
                arkusz1.cell(column=1,row=count_rows).value = lista_naglowek[count_rows-6] + wiersz.text
                arkusz2.cell(column=1,row=count_rows).value = lista_naglowek[count_rows-6] + wiersz.text
            count_rows += 1
            for item in lista_naglowek:
                if wiersz.tag == początek_taga + item:
                    arkusz1.cell(column=1,row=lista_naglowek.index(item)+1).value = item
                    arkusz1.cell(column=3,row=lista_naglowek.index(item)+1).value = wiersz.text
                    arkusz2.cell(column=1,row=lista_naglowek.index(item)+1).value = item
                    arkusz2.cell(column=3,row=lista_naglowek.index(item)+1).value = wiersz.text
            
            if wiersz.tag == początek_taga + 'KodFormularza':
                slownik_atrybutow = wiersz.attrib
        arkusz1.cell(column=1,row=4).value = 'kodSystemowy: ' + slownik_atrybutow['kodSystemowy']
        arkusz1.cell(column=1,row=5).value = 'wersjaSchemy: ' + slownik_atrybutow['wersjaSchemy']
        arkusz2.cell(column=1,row=4).value = 'kodSystemowy: ' + slownik_atrybutow['kodSystemowy']
        arkusz2.cell(column=1,row=5).value = 'wersjaSchemy: ' + slownik_atrybutow['wersjaSchemy']
            
''' PODMIOT '''                

lista_wartości = []
for child in root:
    if child.tag == początek_taga + 'Podmiot1':
        for wiersz in child:
            lista_wartości.append(wiersz.text)
arkusz1.cell(column=1,row=1).value = 'NIP: ' + lista_wartości[0]
arkusz2.cell(column=1,row=1).value = 'NIP: ' + lista_wartości[0]
arkusz1.cell(column=1,row=2).value = 'PelnaNazwa: ' + lista_wartości[1]
arkusz2.cell(column=1,row=2).value = 'PelnaNazwa: ' + lista_wartości[1]

''' SPRZEDAŻ '''

# wstawianie nagłówków kolumn
lista_sprzedaz = ['LpSprzedazy','NrKontrahenta', 'NazwaKontrahenta','AdresKontrahenta',
        'DowodSprzedazy','DataWystawienia','K_15','K_16','K_17','K_18','K_19',
        'K_20','K_21','K_22','K_23','K_24','K_25','K_26','K_27','K_28','K_29','K_30','K_32',
        'K_33','K_34','K_35'] 
list_sales_headers = ['sprzedaż krajowa 5%', 'VAT 5%', 'sprzedaż krajowa 7% lub 8%', 'VAT 7% lub 8%',
     'sprzedaż krajowa 22% lub 23%', 'VAT 22% lub 23%', 'WDT', 'Eksport', "WNT", 'VAT WNT', 
     'Import', 'VAT Import', 'Import uslug z wył. art. 28b', 'VAT imp.usł. z wył. art. 28b', 
     'Import usług art. 28b', "VAT imp. usł. art. 28b", 'Dostawa art. 17 ust. 1 pkt 5', 'VAT dostawa art. 17 ust. 1 pkt 5', 
     'Dostawa art. 17 ust. 1 pkt 7 lub 8 (nabywca)', 'VAT dostawa art. 17 ust. 1 pkt 7 lub 8 (nabywca)'  ]
for item in lista_sprzedaz:
    arkusz1.cell(column=lista_sprzedaz.index(item)+1,row=17).value = item
for item in list_sales_headers:
    arkusz1.cell(column=list_sales_headers.index(item)+7,row=18).value = item

count_sales = 20
for child in root:
    if child.tag == początek_taga + 'SprzedazWiersz':
        for faktura in child:
            for item in lista_sprzedaz:
                if faktura.tag == początek_taga + item:
                    if item[0] == 'K':
                        arkusz1.cell(column=lista_sprzedaz.index(item)+1,row=count_sales).value = round(float(faktura.text),2)
                        arkusz1.cell(column=lista_sprzedaz.index(item)+1,row=count_sales).number_format = '# ##0.00'
                    else:
                        arkusz1.cell(column=lista_sprzedaz.index(item)+1,row=count_sales).value = faktura.text
        count_sales += 1
    
    if child.tag == początek_taga + 'SprzedazCtrl':
        for item in child:
            if item.tag == początek_taga + 'LiczbaWierszySprzedazy':
                liczba_wierszy_sprzedaży = int(item.text)
                arkusz1.cell(column=1,row=13).value = 'Kontrolna liczba wierszy sprzedaży:'
                arkusz1.cell(column=5,row=13).value = liczba_wierszy_sprzedaży
            if item.tag == początek_taga + 'PodatekNalezny':
                podatek_należny = round(float(item.text),2)
                arkusz1.cell(column=1,row=14).value = 'Kontrolna kwota podatku należnego:'
                arkusz1.cell(column=5,row=14).value = podatek_należny
                arkusz1.cell(column=5,row=14).number_format = '# ##0.00'
#konstruowanie formuły excel do komórki J10
sums_sales = str(liczba_wierszy_sprzedaży+19)
literki = ['H', 'J', 'L', 'P', 'R', 'T', 'V', 'X', 'Z']
dluga_formulka = '=sum(E14'
for literka in literki:
    czlon = ') - sum(' + literka + '20:' + literka
    dluga_formulka += czlon
    dluga_formulka += sums_sales
dluga_formulka = dluga_formulka + ')'
arkusz1['A15'] = 'różnica:'
arkusz1['B15'] = dluga_formulka
arkusz1['B15'].number_format = '# ##0.00'

#formuly excel do sumowania kolumn
arkusz1.cell(column=6,row=19).value = 'Razem:'
kolumny = ['G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
for sth in kolumny:    
    formułka = '=sum(' + sth + '20' + ':' + sth + sums_sales + ')'    
    arkusz1[sth + '19'] = formułka    
    arkusz1.cell(column=kolumny.index(sth)+7,row=19).number_format = '# ##0.00'
lista_podatek_nalezny = ['H','J','L','N','P','R','T','V','X','Y','Z']

podatek_należny_razem = 0

print('count_sales',count_sales)


''' ZAKUP '''

lista_zakup = ['LpZakupu','NrDostawcy', 'NazwaDostawcy','AdresDostawcy',
    'DowodZakupu','DataZakupu','K_43','K_44','K_45','K_46'] 

for item in lista_zakup:
    arkusz2.cell(column=lista_zakup.index(item)+1,row=17).value = item

count_purchase = 19
for child in root:
    if child.tag == początek_taga + 'ZakupWiersz':      
        for faktura in child:
            #print(item.tag) 
            
            for item in lista_zakup:
                if faktura.tag == początek_taga + item:
                    if item[0] == 'K':
                        arkusz2.cell(column=lista_zakup.index(item)+1,row=count_purchase).value = round(float(faktura.text),2)
                        arkusz2.cell(column=lista_zakup.index(item)+1,row=count_purchase).number_format = '# ##0.00'
                    else:
                        arkusz2.cell(column=lista_zakup.index(item)+1,row=count_purchase).value = faktura.text
        count_purchase += 1

    if child.tag == początek_taga + 'ZakupCtrl':
        for item in child:
            if item.tag == początek_taga + 'LiczbaWierszyZakupow':
                liczba_wierszy_zakupów = int(item.text)
                arkusz2.cell(column=1,row=13).value = 'Kontrolna liczba wierszy zakupów:'
                arkusz2.cell(column=5,row=13).value = liczba_wierszy_zakupów
            if item.tag == początek_taga + 'PodatekNaliczony':
                podatek_naliczony = round(float(item.text),2)
                arkusz2.cell(column=1,row=14).value = 'Kontrolna kwota podatku naliczonego:'
                arkusz2.cell(column=5,row=14).value = podatek_naliczony
                arkusz2.cell(column=5,row=14).number_format = '# ##0.00'

#formuły excel w wierszu 15
sums_purch = str(liczba_wierszy_zakupów+18)
arkusz2.cell(column=6,row=18).value = 'Razem:'
kolumny = ['G','H','I','J']
for sth in kolumny:
    formułka = '=sum(' + sth + '19' + ':' + sth + sums_purch + ')'
    arkusz2[sth + '18'] = formułka
    arkusz2.cell(column=kolumny.index(sth)+7,row=18).number_format = '# ##0.00'

#formuła excel różnicy:
arkusz2['A15'] = 'różnica:'
arkusz2['B15'] = '=sum(E14)-sum(H19:H' + sums_purch + ')-sum(J19:J' + sums_purch + ')'
arkusz2['B15'].number_format = '# ##0.00'
print('count_purchase',count_purchase)

nazwa_pliku = "JPK_" + od + "_" + do + "_v_" + wersja_jpk + ".xlsx" 
raport.save(nazwa_pliku) 

# umożliwia otwarcie pliku xlsx z Windows (TKinter+os):
os.startfile(askopenfilename())
